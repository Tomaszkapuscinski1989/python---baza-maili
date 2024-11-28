import xlsxwriter
import openpyxl
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from tkinter import filedialog
import os
import datetime

import contextManager


class Exel:

    def __init__(self):
        self.nazwaBazy = ""
        self.nazwaPliku = ""

    def export_do_exela(self):
        try:
            self.nazwaPliku = filedialog.asksaveasfilename(
                defaultextension="*.xlsx", filetypes=(("xlsx", "*.xlsx"),)
            )

            workbook = xlsxwriter.Workbook(self.nazwaPliku)
            worksheet = workbook.add_worksheet()

            with contextManager.open_base(self.nazwaBazy) as c:
                c.execute("select *, oid FROM klienci")
                r = c.fetchall()

            Site_header_format = workbook.add_format(
                {
                    "bold": True,
                    "font_color": "black",
                    "align": "center",
                    "valign": "vcenter",
                    "font_size": 30,
                }
            )

            header_format = workbook.add_format(
                {
                    "bold": True,
                    "font_color": "white",
                    "align": "center",
                    "valign": "vcenter",
                    "bg_color": "black",
                }
            )
            dane_foemat1 = workbook.add_format(
                {
                    "font_color": "white",
                    "bg_color": "gray",
                    "align": "center",
                    "right": 1,
                }
            )
            dane_foemat2 = workbook.add_format(
                {
                    "font_color": "black",
                    "bg_color": "white",
                    "align": "center",
                    "right": 1,
                }
            )
            col = 0
            exelHeader = [
                {"row": 3, "col": 0, "text": "Id", "format": header_format},
                {"row": 3, "col": 1, "text": "Imię", "format": header_format},
                {"row": 3, "col": 2, "text": "Nazwisko", "format": header_format},
                {"row": 3, "col": 3, "text": "Telefon", "format": header_format},
                {"row": 3, "col": 4, "text": "Nazwa firmy", "format": header_format},
                {
                    "row": 3,
                    "col": 5,
                    "text": "Strona internetowa",
                    "format": header_format,
                },
                {"row": 3, "col": 6, "text": "Adres e-mail:", "format": header_format},
            ]
            for item in exelHeader:
                worksheet.write(item["row"], item["col"], item["text"], item["format"])

            col = 0
            for index, item in enumerate(r, start=4):
                if index % 2 == 0:
                    format = dane_foemat1
                else:
                    format = dane_foemat2
                worksheet.write(index, col, item[-1], format)
                worksheet.write(index, col + 1, item[0], format)
                worksheet.write(index, col + 2, item[1], format)
                worksheet.write(index, col + 3, item[2], format)
                worksheet.write(index, col + 4, item[3], format)
                worksheet.write(index, col + 5, item[4], format)
                worksheet.write(index, col + 6, item[5], format)

            worksheet.set_column("A:A", 5)
            worksheet.set_column("B:D", 15)
            worksheet.set_column("E:F", 20)
            worksheet.set_column("G:G", 40)

            worksheet.set_row(0, 80)
            worksheet.set_row(1, 40)
            worksheet.set_row(3, 30)
            worksheet.merge_range("A1:G1", f"Spis danych klieniów", Site_header_format)
            worksheet.merge_range("A2:D2", "Data wykonania:", Site_header_format)
            now = datetime.datetime.now()
            worksheet.merge_range("E2:G2", f"{now:%d.%m.%Y %-H:%M}", Site_header_format)
            workbook.close()
        except FileNotFoundError:
            messagebox.showerror("Error", "Nie Podano nazwy pliku")
        except xlsxwriter.exceptions.FileCreateError:
            messagebox.showerror("Error", "Nie Podano nazwy pliku")
        except Exception:
            messagebox.showerror("Error", "Coś poszło nie tak")
        else:
            messagebox.showinfo("info", "Dane zostały poprawnie wyeksportowane")

    def import_z_exela(self):
        try:
            self.nazwaPliku = filedialog.askopenfilename(
                filetypes=(
                    ("xlsx", "*.xlsx"),
                    ("xlsm", "*.xlsm"),
                    ("xltx", "*.xltx"),
                    ("xltm", "*.xltm"),
                )
            )

            dataframe = openpyxl.load_workbook(self.nazwaPliku)

            dataframe1 = dataframe.active

            with contextManager.open_base(self.nazwaBazy) as c:
                c.execute("DELETE from klienci")

            for row in range(4, dataframe1.max_row):
                g = []
                for col in dataframe1.iter_cols(1, dataframe1.max_column):

                    g.append(col[row].value)

                with contextManager.open_base(self.nazwaBazy) as c:
                    c.execute(
                        " INSERT INTO klienci VALUES (:imie, :nazwisko, :telefon, :nazaFirmy, :stronaInternetowa, :adresMail)",
                        {
                            "imie": g[1],
                            "nazwisko": g[2],
                            "telefon": int(g[3]),
                            "nazaFirmy": g[4],
                            "stronaInternetowa": g[5],
                            "adresMail": g[6],
                        },
                    )
        except openpyxl.utils.exceptions.InvalidFileException:
            messagebox.showerror("Error", "Nie wybrano pliku")
        except Exception:
            messagebox.showerror("Error", "Coś poszło nie tak")
        else:
            messagebox.showinfo("info", "Dane zostały poprawnie zainportowane")
